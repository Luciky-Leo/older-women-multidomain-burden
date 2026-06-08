from __future__ import annotations

import shutil
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUTS = PROJECT_ROOT / "outputs"
PHASE47_GMM_DIR = OUTPUTS / "phase47_gmm_bootstrap_200"
PACKAGE_DIR = PROJECT_ROOT / "manuscript" / "bmc_geriatrics_submission_ready_20260605"

COHORTS = ["CHARLS", "ELSA", "HRS", "KLoSA", "LASI", "MHAS", "SHARE"]


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, low_memory=False, encoding="utf-8-sig")


def write_csv(df: pd.DataFrame, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def evidence_sentence_for_gmm(row: pd.Series) -> str:
    near_singular = int(row.get("any_near_singular_covariance", 0) or 0) == 1
    median_ari = row.get("median_ari_vs_reference")
    p10_ari = row.get("p10_ari_vs_reference")
    if near_singular:
        return "near-singular covariance persists; descriptive partition only"
    if pd.notna(median_ari) and pd.notna(p10_ari) and float(median_ari) >= 0.75 and float(p10_ari) >= 0.60:
        return "bootstrap agreement acceptable but still interpreted with harmonization guardrails"
    return "bootstrap agreement below stability threshold; descriptive partition only"


def build_gmm_tables() -> dict[str, Path]:
    extended_summary = PHASE47_GMM_DIR / "phase32_gmm_stability_summary.csv"
    extended_bootstrap = PHASE47_GMM_DIR / "phase32_gmm_bootstrap_stability.csv"
    extended_covariance = PHASE47_GMM_DIR / "phase32_gmm_covariance_diagnostics.csv"

    if not extended_summary.exists():
        raise FileNotFoundError(
            f"Extended GMM bootstrap summary is missing: {extended_summary}. "
            "Finish scripts/build_phase32_gmm_stability_diagnostics.py before running Phase47."
        )

    summary = read_csv(extended_summary)
    summary["bootstrap_replicates"] = pd.to_numeric(summary["bootstrap_replicates"], errors="coerce").astype("Int64")
    summary["phase47_reviewer_reading"] = summary.apply(evidence_sentence_for_gmm, axis=1)
    summary["source_note"] = "Extended nonparametric bootstrap refits from original complete-domain rows; no simulated participant data."
    summary_out = write_csv(summary, OUTPUTS / "phase47_gmm_bootstrap_robustness_summary.csv")

    paths = {"gmm_summary": summary_out}
    if extended_bootstrap.exists():
        bootstrap = read_csv(extended_bootstrap)
        bootstrap["source_note"] = "Extended nonparametric bootstrap refits from original complete-domain rows."
        paths["gmm_replicates"] = write_csv(bootstrap, OUTPUTS / "phase47_gmm_bootstrap_robustness_replicates.csv")
    if extended_covariance.exists():
        paths["gmm_covariance"] = shutil.copy2(
            extended_covariance,
            OUTPUTS / "phase47_gmm_covariance_diagnostics_extended.csv",
        )
    return paths


def build_sex_interaction_review() -> Path:
    summary = read_csv(OUTPUTS / "phase41_all_sex_lfo_sex_interaction_summary.csv")
    summary["interaction_p_value"] = pd.to_numeric(summary["interaction_p_value"], errors="coerce")
    summary["nominal_interaction_p_lt_0_05"] = summary["interaction_p_value"].lt(0.05).astype(int)
    summary["phase47_interpretation"] = summary["nominal_interaction_p_lt_0_05"].map(
        {
            1: "nominal cohort-specific interaction; not consistent across cohorts",
            0: "no nominal interaction signal",
        }
    )
    summary["women_only_implication"] = (
        "Supports focused older-women reporting only; does not establish a women-specific mechanism."
    )
    keep_cols = [
        "analysis_tier",
        "cohort",
        "n",
        "events",
        "event_pct",
        "female_n",
        "male_n",
        "interaction_or",
        "interaction_ci",
        "interaction_p_value",
        "nominal_interaction_p_lt_0_05",
        "phase47_interpretation",
        "scale_note",
        "women_only_implication",
    ]
    return write_csv(summary[keep_cols], OUTPUTS / "phase47_women_only_sex_interaction_summary_for_review.csv")


def build_survey_review() -> Path:
    survey = read_csv(OUTPUTS / "phase41_survey_design_triplet_status.csv")
    for col in [
        "weight_cleaned_candidate",
        "weight_metadata_evidence",
        "psu_cleaned_candidate",
        "psu_metadata_evidence",
        "strata_cleaned_candidate",
        "strata_metadata_evidence",
    ]:
        survey[col] = pd.to_numeric(survey[col], errors="coerce").fillna(0).astype(int)
    survey["candidate_triplet_all_three_cleaned_headers"] = (
        (survey["weight_cleaned_candidate"] == 1)
        & (survey["psu_cleaned_candidate"] == 1)
        & (survey["strata_cleaned_candidate"] == 1)
    ).astype(int)
    survey["candidate_triplet_all_three_any_metadata"] = (
        (survey["weight_metadata_evidence"] == 1)
        & (survey["psu_metadata_evidence"] == 1)
        & (survey["strata_metadata_evidence"] == 1)
    ).astype(int)
    survey["phase47_population_claim_implication"] = survey["harmonized_triplet_status"].map(
        lambda status: (
            "not sufficient for harmonized seven-cohort survey-weighted prevalence claim"
            if "no_codebook" in str(status)
            else "candidate triplet needs manual codebook confirmation before weighted sensitivity"
        )
    )
    keep_cols = [
        "cohort",
        "weight_cleaned_candidate",
        "psu_cleaned_candidate",
        "strata_cleaned_candidate",
        "candidate_triplet_all_three_cleaned_headers",
        "candidate_triplet_all_three_any_metadata",
        "harmonized_triplet_status",
        "analysis_decision",
        "phase47_population_claim_implication",
    ]
    return write_csv(survey[keep_cols], OUTPUTS / "phase47_survey_design_review_summary.csv")


def build_endpoint_feasibility() -> Path:
    mortality = read_csv(OUTPUTS / "phase41_mortality_secondary_formal_model_table.csv")
    hospitalization = read_csv(OUTPUTS / "phase41_hospitalization_candidate_status.csv")

    rows: list[dict[str, object]] = []
    for cohort in COHORTS:
        mort_row = mortality[mortality["cohort"] == cohort]
        if mort_row.empty:
            rows.append(
                {
                    "cohort": cohort,
                    "endpoint": "all_cause_mortality",
                    "status": "unavailable_current_cleaned_pass",
                    "n": pd.NA,
                    "events": pd.NA,
                    "event_pct": pd.NA,
                    "evidence_level": "not modelled",
                    "manuscript_use": "unavailable",
                }
            )
        else:
            item = mort_row.iloc[0]
            modelled = str(item.get("analysis_status", "")).lower().startswith("formal")
            rows.append(
                {
                    "cohort": cohort,
                    "endpoint": "all_cause_mortality",
                    "status": item.get("analysis_status"),
                    "n": item.get("n"),
                    "events": item.get("deaths"),
                    "event_pct": item.get("event_pct"),
                    "evidence_level": "formal secondary Cox guardrail" if modelled else "unavailable",
                    "manuscript_use": "secondary non-circular guardrail" if modelled else "outside validation denominator",
                }
            )

        hosp_row = hospitalization[hospitalization["cohort"] == cohort]
        if hosp_row.empty:
            status = "not modelable in current cleaned pass"
            variables = ""
            n = events = event_pct = pd.NA
        else:
            item = hosp_row.iloc[0]
            status = item.get("status")
            variables = item.get("variables_used")
            n = item.get("n")
            events = item.get("events")
            event_pct = item.get("event_pct")
        rows.append(
            {
                "cohort": cohort,
                "endpoint": "hospitalization",
                "status": status,
                "n": n,
                "events": events,
                "event_pct": event_pct,
                "evidence_level": "header-only candidate model" if status == "candidate modelable" else "not modelled",
                "variables_used": variables,
                "manuscript_use": "sensitivity feasibility only until codebook mapping" if status == "candidate modelable" else "unavailable",
            }
        )

        for endpoint in ["institutionalization", "care_dependence"]:
            rows.append(
                {
                    "cohort": cohort,
                    "endpoint": endpoint,
                    "status": "unavailable_current_cleaned_pass",
                    "n": pd.NA,
                    "events": pd.NA,
                    "event_pct": pd.NA,
                    "evidence_level": "not modelled",
                    "manuscript_use": "endpoint availability matrix only; no inferential claim",
                }
            )

    matrix = pd.DataFrame(rows)
    return write_csv(matrix, OUTPUTS / "phase47_hard_endpoint_feasibility_matrix.csv")


def build_guardrail_summary(paths: dict[str, Path]) -> Path:
    gmm = read_csv(paths["gmm_summary"])
    sex = read_csv(paths["sex_review"])
    survey = read_csv(paths["survey_review"])
    endpoint = read_csv(paths["endpoint_matrix"])

    gmm_rep = int(pd.to_numeric(gmm["bootstrap_replicates"], errors="coerce").max())
    nominal_sex = int(pd.to_numeric(sex["nominal_interaction_p_lt_0_05"], errors="coerce").sum())
    candidate_triplets = int(
        (survey["harmonized_triplet_status"] == "candidate_triplet_needs_manual_confirmation").sum()
    )
    mortality_modelled = int(
        (
            (endpoint["endpoint"] == "all_cause_mortality")
            & endpoint["evidence_level"].eq("formal secondary Cox guardrail")
        ).sum()
    )
    hospitalization_candidate = int(
        (
            (endpoint["endpoint"] == "hospitalization")
            & endpoint["evidence_level"].eq("header-only candidate model")
        ).sum()
    )

    rows = [
        {
            "reviewer_risk_area": "GMM stability",
            "phase47_enhancement": f"Extended nonparametric bootstrap to {gmm_rep} refits per selected solution.",
            "key_result": "Near-singular covariance guardrail remained present for selected full-covariance solutions.",
            "manuscript_implication": "Keep profile labels descriptive; do not claim stable latent endotypes.",
        },
        {
            "reviewer_risk_area": "Women-only scope",
            "phase47_enhancement": "All-sex LFO severity-by-sex interaction review table.",
            "key_result": f"{nominal_sex} of 6 modelled cohorts had nominal interaction p<0.05.",
            "manuscript_implication": "Use focused older-women framing, not women-specific mechanism language.",
        },
        {
            "reviewer_risk_area": "Survey design",
            "phase47_enhancement": "Cleaned-header plus metadata triplet availability summary.",
            "key_result": f"{candidate_triplets} cohorts had candidate triplets needing manual confirmation; no seven-cohort confirmed triplet.",
            "manuscript_implication": "No survey-weighted prevalence or population estimate claim.",
        },
        {
            "reviewer_risk_area": "Hard endpoints",
            "phase47_enhancement": "Endpoint feasibility matrix across mortality, hospitalization, institutionalization and care-dependence.",
            "key_result": f"Mortality formal secondary models in {mortality_modelled} cohorts; hospitalization candidate models in {hospitalization_candidate} cohorts; institutionalization/care-dependence unavailable.",
            "manuscript_implication": "Mortality remains secondary; hospitalization remains feasibility/sensitivity; no independent hard-outcome validation claim.",
        },
    ]
    return write_csv(pd.DataFrame(rows), OUTPUTS / "phase47_reviewer_risk_enhancement_summary.csv")


def add_dataframe_sheet(workbook_path: Path, sheet_name: str, df: pd.DataFrame) -> None:
    wb = load_workbook(workbook_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    rows = [list(df.columns)] + df.fillna("").astype(str).values.tolist()
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 90))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(10, min(max_len + 2, 55))
    if "README" in wb.sheetnames:
        readme = wb["README"]
        next_row = readme.max_row + 1
        readme.cell(next_row, 1).value = f"Phase47 sheet added: {sheet_name}"
        readme.cell(next_row, 2).value = datetime.now(timezone.utc).isoformat()
    wb.save(workbook_path)


def update_submission_workbooks(paths: dict[str, Path]) -> None:
    add_dataframe_sheet(
        PACKAGE_DIR / "additional_file_2_profile_stability_and_descriptive_profiles.xlsx",
        "gmm_bootstrap_ext",
        read_csv(paths["gmm_summary"]),
    )
    if "gmm_replicates" in paths:
        add_dataframe_sheet(
            PACKAGE_DIR / "additional_file_2_profile_stability_and_descriptive_profiles.xlsx",
            "gmm_replicates_ext",
            read_csv(paths["gmm_replicates"]),
        )
    add_dataframe_sheet(
        PACKAGE_DIR / "additional_file_4_secondary_endpoints_and_prediction_guardrails.xlsx",
        "endpoint_feasibility",
        read_csv(paths["endpoint_matrix"]),
    )
    add_dataframe_sheet(
        PACKAGE_DIR / "additional_file_5_survey_design_and_sex_comparator_audits.xlsx",
        "sex_interaction_review",
        read_csv(paths["sex_review"]),
    )
    add_dataframe_sheet(
        PACKAGE_DIR / "additional_file_5_survey_design_and_sex_comparator_audits.xlsx",
        "survey_design_review",
        read_csv(paths["survey_review"]),
    )
    add_dataframe_sheet(
        PACKAGE_DIR / "additional_file_5_survey_design_and_sex_comparator_audits.xlsx",
        "reviewer_risk_summary",
        read_csv(paths["guardrail_summary"]),
    )


def main() -> None:
    paths = build_gmm_tables()
    paths["sex_review"] = build_sex_interaction_review()
    paths["survey_review"] = build_survey_review()
    paths["endpoint_matrix"] = build_endpoint_feasibility()
    paths["guardrail_summary"] = build_guardrail_summary(paths)
    update_submission_workbooks(paths)

    print("Phase47 reviewer-risk enhancement complete.")
    for key, path in paths.items():
        print(f"{key}: {path}")


if __name__ == "__main__":
    main()
