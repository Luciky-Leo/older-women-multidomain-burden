from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PACKAGE_DIR = PROJECT_ROOT / "manuscript" / "bmc_geriatrics_submission_burden_profiles_rescue"
OUT_DIR = PACKAGE_DIR / "submission_workbooks"


WORKBOOKS = [
    {
        "file_name": "additional_file_1_harmonization_and_cohort_construction.xlsx",
        "title": "Cohort construction, harmonization and baseline covariate metadata",
        "purpose": "Item-level harmonization, cohort denominator/tier lock, harmonization risk and baseline covariate availability.",
        "sheets": [
            ("item_crosswalk", "additional_file_1_item_level_harmonization_crosswalk.csv"),
            ("cohort_tier_lock", "additional_file_2_cohort_tier_lock.csv"),
            ("harmonization_risk_matrix", "additional_file_9_harmonization_risk_matrix.csv"),
            ("baseline_covariates", "additional_file_12_baseline_clinical_design_covariate_availability.csv"),
        ],
    },
    {
        "file_name": "additional_file_2_profile_stability_and_descriptive_profiles.xlsx",
        "title": "Descriptive profile and model-stability guardrails",
        "purpose": "GMM selection/stability outputs, covariance diagnostics, descriptive class dictionaries and sensitivity-cohort profile support.",
        "sheets": [
            ("gmm_stability_summary", "additional_file_4_gmm_stability_summary.csv"),
            ("gmm_covariance", "additional_file_5_gmm_covariance_diagnostics.csv"),
            ("class_dictionary", "additional_file_7_selected_class_dictionary.csv"),
            ("profile_summary", "additional_file_8_profile_family_summary.csv"),
            ("profile_guardrails", "additional_file_13_profile_stability_guardrails.csv"),
            ("algorithm_robustness", "additional_file_17_gmm_algorithm_robustness.csv"),
            ("sensitivity_profiles", "additional_file_25_sensitivity_cohort_profile_support.csv"),
        ],
    },
    {
        "file_name": "additional_file_3_lfo_functional_change_associations.xlsx",
        "title": "Leave-functional-domain-out functional-change association outputs",
        "purpose": "Decoupled LFO functional-change models, leakage audit, class risks, AUC intervals and functional-change heterogeneity summaries.",
        "sheets": [
            ("decoupled_comparison", "additional_file_3_decoupled_validation_comparison.csv"),
            ("leakage_audit", "additional_file_6_functional_endpoint_leakage_audit.csv"),
            ("strict_core_lfo", "additional_file_14_strict_core_lfo_functional_change_association.csv"),
            ("lfo_sensitivity", "additional_file_15_lfo_sensitivity_rows_removed_from_main.csv"),
            ("class_risks", "additional_file_16_functional_association_class_risks.csv"),
            ("auc_bootstrap", "additional_file_18_auc_bootstrap_intervals.csv"),
            ("heterogeneity", "additional_file_32_cross_cohort_heterogeneity_summary.csv"),
        ],
    },
    {
        "file_name": "additional_file_4_secondary_endpoints_and_prediction_guardrails.xlsx",
        "title": "Secondary endpoint and prediction guardrail outputs",
        "purpose": "Formal mortality guardrails, hospitalization candidate analyses, calibration metrics, decision-curve summaries and prediction guardrails.",
        "sheets": [
            ("mortality_guardrail", "additional_file_23_mortality_secondary_guardrail_table.csv"),
            ("mortality_formal", "additional_file_26_phase41_mortality_secondary_formal_model_table.csv"),
            ("hospitalization_models", "additional_file_27_phase41_hospitalization_candidate_model_comparison.csv"),
            ("hospitalization_status", "additional_file_27b_phase41_hospitalization_candidate_status.csv"),
            ("calibration", "additional_file_30_phase41_calibration_metrics.csv"),
            ("decision_curve", "additional_file_30b_phase41_decision_curve_summary.csv"),
            ("prediction_guardrails", "additional_file_31_phase41_secondary_endpoint_prediction_guardrail_summary.csv"),
            ("heterogeneity", "additional_file_32_cross_cohort_heterogeneity_summary.csv"),
        ],
    },
    {
        "file_name": "additional_file_5_survey_design_and_sex_comparator_audits.xlsx",
        "title": "Survey-design and sex-comparator audits",
        "purpose": "Survey-design variable/codebook audits, survey triplet status, baseline male-comparator summaries and all-sex LFO interaction terms.",
        "sheets": [
            ("survey_variable_audit", "additional_file_19_survey_design_variable_audit.csv"),
            ("survey_raw_codebook", "additional_file_28_phase41_survey_design_raw_codebook_audit.csv"),
            ("survey_triplet_status", "additional_file_28b_phase41_survey_design_triplet_status.csv"),
            ("male_summary", "additional_file_20_baseline_male_comparator_domain_summary.csv"),
            ("male_contrasts", "additional_file_21_baseline_male_comparator_domain_contrasts.csv"),
            ("sex_interaction_summary", "additional_file_29_phase41_all_sex_lfo_sex_interaction_summary.csv"),
            ("sex_interaction_terms", "additional_file_29b_phase41_all_sex_lfo_sex_interaction_terms.csv"),
        ],
    },
]


def read_csv_rows(file_name: str) -> list[list[str]]:
    source = PACKAGE_DIR / file_name
    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.reader(handle))


def add_rows(ws, rows: list[list[str]]) -> None:
    for row in rows:
        ws.append(row)
    if rows:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(10, min(max_len + 2, 60))


def build_workbook(spec: dict) -> Path:
    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    add_rows(
        readme,
        [
            ["Field", "Value"],
            ["Workbook title", spec["title"]],
            ["Purpose", spec["purpose"]],
            ["Source note", "Generated from project CSV outputs produced from approved downloaded cohort files cleaned by the authors."],
            ["Generated by", "scripts/build_bmc_submission_workbooks.py"],
            ["Generated on", datetime.now(timezone.utc).isoformat()],
            ["Included sheets", "; ".join(sheet_name for sheet_name, _ in spec["sheets"])],
        ],
    )

    for sheet_name, csv_file in spec["sheets"]:
        rows = read_csv_rows(csv_file)
        ws = wb.create_sheet(title=sheet_name)
        add_rows(ws, rows)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / spec["file_name"]
    wb.save(out_path)
    return out_path


def main() -> None:
    for spec in WORKBOOKS:
        out_path = build_workbook(spec)
        print(out_path)


if __name__ == "__main__":
    main()
