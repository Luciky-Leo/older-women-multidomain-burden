from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
PKG = ROOT / "manuscript" / "bmc_womens_health_submission_ready_20260608"
OUT = ROOT / "outputs"


WORKBOOKS = [
    (
        "additional_file_1_harmonization_and_cohort_construction.xlsx",
        {
            "README": "Workbook overview.",
            "item_crosswalk": "Item-level cohort-domain harmonization crosswalk.",
            "cohort_tier_lock": "Locked evidence-tier and analysis-role definitions.",
            "harmonization_risk_matrix": "Cohort-by-domain measurement risk matrix.",
            "baseline_covariates": "Baseline clinical, design and covariate availability metadata.",
            "age_distribution": "Women aged 50 years or older split into 50-64 and 65+ strata by cohort.",
        },
    ),
    (
        "additional_file_2_profile_stability_and_descriptive_profiles.xlsx",
        {
            "README": "Workbook overview.",
            "gmm_stability_summary": "Selected GMM stability summaries.",
            "gmm_covariance": "Covariance diagnostics for selected GMM solutions.",
            "class_dictionary": "Descriptive class labels and domain signatures.",
            "profile_summary": "Descriptive profile distribution summaries.",
            "profile_guardrails": "Profile interpretation guardrail table.",
            "algorithm_robustness": "Alternative algorithm agreement summaries.",
            "sensitivity_profiles": "Sensitivity-tier descriptive profile support.",
            "gmm_bootstrap_ext": "Extended 200-refit bootstrap robustness summary.",
            "gmm_replicates_ext": "Bootstrap replicate-level stability outputs.",
        },
    ),
    (
        "additional_file_3_lfo_functional_change_associations.xlsx",
        {
            "README": "Workbook overview.",
            "decoupled_comparison": "LFO profile versus continuous-score comparator outputs.",
            "leakage_audit": "Functional endpoint leakage and LFO design audit.",
            "strict_core_lfo": "Strict-core LFO functional-change association table.",
            "lfo_sensitivity": "SHARE/KLoSA sensitivity rows.",
            "class_risks": "Class-level functional-change risks.",
            "auc_bootstrap": "AUC bootstrap intervals and deltas.",
            "heterogeneity": "Cross-cohort functional-change heterogeneity summaries.",
        },
    ),
    (
        "additional_file_4_secondary_endpoints_and_prediction_guardrails.xlsx",
        {
            "README": "Workbook overview.",
            "mortality_guardrail": "Secondary mortality guardrail table.",
            "mortality_formal": "Formal secondary Cox model outputs.",
            "hospitalization_models": "Header-available hospitalization candidate models.",
            "hospitalization_status": "Hospitalization endpoint status by cohort.",
            "calibration": "Apparent in-sample calibration diagnostics.",
            "decision_curve": "Apparent decision-curve summaries.",
            "prediction_guardrails": "Prediction interpretation guardrails.",
            "heterogeneity": "Cross-cohort mortality heterogeneity summaries.",
            "endpoint_feasibility": "Hard-endpoint feasibility matrix.",
        },
    ),
    (
        "additional_file_5_survey_design_and_sex_comparator_audits.xlsx",
        {
            "README": "Workbook overview.",
            "survey_variable_audit": "Cleaned-header survey-design variable audit.",
            "survey_raw_codebook": "Raw/codebook/dofile survey-design audit evidence.",
            "survey_triplet_status": "Weight, PSU and strata triplet availability status.",
            "male_summary": "All-sex baseline male/female summary.",
            "male_contrasts": "All-sex baseline domain contrasts.",
            "sex_interaction_summary": "All-sex LFO severity-by-sex interaction summary.",
            "sex_interaction_terms": "Interaction model terms.",
            "sex_interaction_review": "Women-only rationale and sex-interaction review table.",
            "survey_design_review": "Survey-design limitations and interpretation review.",
            "reviewer_risk_summary": "Reviewer-risk response summary for unresolved guardrails.",
        },
    ),
]


def autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 72)


def write_dataframe_sheet(wb, name: str, df: pd.DataFrame, index: int | None = None) -> None:
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, index=index)
    for j, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for i, row in enumerate(df.itertuples(index=False), 2):
        for j, value in enumerate(row, 1):
            if pd.isna(value):
                value = None
            ws.cell(row=i, column=j, value=value)
    autosize(ws)


def build_age_distribution() -> pd.DataFrame:
    src = pd.read_csv(OUT / "phase58_midlife_age_distribution.csv")
    keep = [
        "cohort",
        "source_age_nonmissing_n",
        "source_age_50_64_n",
        "source_age_50_64_pct",
        "source_age_ge65_n",
        "source_age_ge65_pct",
        "complete_age_nonmissing_n",
        "complete_age_50_64_n",
        "complete_age_50_64_pct",
        "complete_age_ge65_n",
        "complete_age_ge65_pct",
    ]
    out = src[keep].copy()
    for col in out.columns:
        if col.endswith("_pct"):
            out[col] = pd.to_numeric(out[col], errors="coerce").round(1)
    return out


def main() -> None:
    index_rows = []
    age_df = build_age_distribution()

    for workbook_name, descriptions in WORKBOOKS:
        path = PKG / workbook_name
        wb = load_workbook(path)

        if workbook_name.startswith("additional_file_1"):
            write_dataframe_sheet(wb, "age_distribution", age_df)

        sheet_names = [s for s in wb.sheetnames if s != "sheet_index"]
        index_df = pd.DataFrame(
            {
                "sheet_name": sheet_names,
                "description": [descriptions.get(s, "Data table used by the manuscript package.") for s in sheet_names],
            }
        )
        write_dataframe_sheet(wb, "sheet_index", index_df, index=0)
        wb.save(path)

        for _, row in index_df.iterrows():
            index_rows.append(
                {
                    "workbook": workbook_name,
                    "sheet_name": row["sheet_name"],
                    "description": row["description"],
                }
            )

    index = pd.DataFrame(index_rows)
    index.to_csv(OUT / "phase58_workbook_sheet_index.csv", index=False, encoding="utf-8-sig")
    report = [
        "# Phase 58 BMC Women's Health Review Fixes",
        "",
        "- Added `age_distribution` to Additional file 1 using `outputs/phase58_midlife_age_distribution.csv`.",
        "- Added a `sheet_index` tab to Additional files 1-5.",
        "- Sheet-index manifest: `outputs/phase58_workbook_sheet_index.csv`.",
        "- Table 1 now reports the complete-domain 50-64/65+ split by cohort.",
    ]
    (OUT / "phase58_bmc_womens_health_review_fixes.md").write_text("\n".join(report) + "\n", encoding="utf-8")
    print(OUT / "phase58_bmc_womens_health_review_fixes.md")


if __name__ == "__main__":
    main()
