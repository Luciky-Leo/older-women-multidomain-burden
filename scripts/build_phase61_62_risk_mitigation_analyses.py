from __future__ import annotations

import argparse
import math
import re
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.api as sm
import statsmodels.formula.api as smf
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from sklearn.cluster import KMeans
from sklearn.metrics import adjusted_rand_score, roc_auc_score, silhouette_score
from sklearn.mixture import GaussianMixture


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
PKG = ROOT / "manuscript" / "bmc_womens_health_submission_ready_20260608"

COHORT_ORDER = ["CHARLS", "ELSA", "HRS", "KLoSA", "LASI", "MHAS", "SHARE"]
STRICT_LFO_COHORTS = ["CHARLS", "ELSA", "HRS", "MHAS"]
DOMAIN_COLS = [
    "functional_score",
    "cognitive_score",
    "affective_score",
    "cardiometabolic_chronic_score",
]
LFO_PROFILE_COLS = ["cognitive_score", "affective_score", "cardiometabolic_chronic_score"]
PHASE62_BOOTSTRAP_REPLICATES = 20
RANDOM_SEED = 20260608
PHASE62_BOOTSTRAP_METHODS = {"gmm_diag", "kmeans", "severity_quantile_k"}
PHASE62_FULL_FIT_GMM_N_INIT = 5
PHASE62_BOOTSTRAP_GMM_N_INIT = 1
PHASE62_FULL_FIT_KMEANS_N_INIT = 20
PHASE62_BOOTSTRAP_KMEANS_N_INIT = 5
PHASE62_BOOTSTRAP_MAX_N = 3000

STABLE_MEDIAN_ARI = 0.90
STABLE_P10_ARI = 0.75
MIN_CLASS_PCT_GATE = 5.0
MIN_GMM_VAR = 1e-5
MAX_GMM_CONDITION = 1e6


def read_csv(name: str, **kwargs) -> pd.DataFrame:
    return pd.read_csv(OUT / name, low_memory=False, **kwargs)


def write_csv(df: pd.DataFrame, name: str) -> Path:
    path = OUT / name
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def normalize_wave(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        return text[:-2]
    return text


def keyify(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in ["participant_id", "wave"]:
        if col in out.columns:
            out[col] = out[col].astype(str).map(normalize_wave)
    return out


def coerce_numeric(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    out = df.copy()
    for column in columns:
        if column in out.columns:
            out[column] = pd.to_numeric(out[column], errors="coerce")
    return out


def safe_pct(num: float, den: float) -> float:
    if not den or pd.isna(den):
        return np.nan
    return float(num) / float(den) * 100.0


def safe_exp(value: float) -> float:
    if pd.isna(value):
        return np.nan
    if value > 700:
        return math.inf
    if value < -700:
        return 0.0
    return float(math.exp(value))


def weighted_mean(values: pd.Series, weights: pd.Series) -> float:
    mask = values.notna() & weights.notna() & (weights > 0)
    if not mask.any():
        return np.nan
    return float(np.average(values[mask].astype(float), weights=weights[mask].astype(float)))


def var_id(header: str) -> str:
    return str(header).strip().strip('"').split(" ", 1)[0].strip()


def read_header_map(path: Path) -> dict[str, str]:
    import csv

    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        raw_header = next(csv.reader(handle))
    return {var_id(raw): raw for raw in raw_header}


def find_clean_csv_from_audit(cohort: str) -> Path | None:
    audit = read_csv("phase38_survey_design_variable_audit.csv")
    paths = audit.loc[audit["cohort"].eq(cohort), "source_file"].dropna().astype(str).unique()
    if not len(paths):
        return None
    path = Path(paths[0])
    return path if path.exists() else None


def cleaned_survey_candidates() -> pd.DataFrame:
    cleaned = read_csv("phase38_survey_design_variable_audit.csv")
    cleaned["variable"] = cleaned["variable"].astype("string")
    rows = []
    for cohort in COHORT_ORDER:
        for concept in ["weight", "psu", "strata"]:
            g = cleaned[(cleaned["cohort"].eq(cohort)) & (cleaned["concept"].eq(concept))].copy()
            nonempty = g[
                g["variable"].notna()
                & g["variable"].astype(str).str.strip().ne("")
                & ~g["harmonization_status"].astype(str).str.contains("no_cleaned_candidate", na=False)
            ]
            rows.append(
                {
                    "cohort": cohort,
                    "concept": concept,
                    "cleaned_variable_count": int(nonempty["variable"].nunique()),
                    "cleaned_variables": ";".join(sorted(nonempty["variable"].astype(str).unique())),
                    "cleaned_status": "available_in_cleaned_csv" if not nonempty.empty else "not_available_in_cleaned_csv",
                }
            )
    return pd.DataFrame(rows)


def corrected_survey_design_status() -> pd.DataFrame:
    cleaned = cleaned_survey_candidates()
    raw = read_csv("phase41_survey_design_raw_codebook_audit.csv")
    raw["variable_or_line"] = raw["variable_or_line"].astype("string")
    raw_nonempty = raw[
        raw["variable_or_line"].notna()
        & raw["variable_or_line"].astype(str).str.strip().ne("")
        & raw["source_type"].astype(str).ne("cleaned_header")
    ].copy()

    rows = []
    for cohort in COHORT_ORDER:
        row: dict[str, object] = {"cohort": cohort}
        for concept in ["weight", "psu", "strata"]:
            c = cleaned[(cleaned["cohort"].eq(cohort)) & (cleaned["concept"].eq(concept))]
            r = raw_nonempty[(raw_nonempty["cohort"].eq(cohort)) & (raw_nonempty["concept"].eq(concept))]
            row[f"{concept}_cleaned_variable_count"] = int(c["cleaned_variable_count"].iloc[0]) if not c.empty else 0
            row[f"{concept}_cleaned_variables"] = c["cleaned_variables"].iloc[0] if not c.empty else ""
            row[f"{concept}_metadata_nonempty_mentions"] = int(len(r))
            row[f"{concept}_metadata_evidence"] = int(len(r) > 0)

        full_cleaned_triplet = all(int(row[f"{c}_cleaned_variable_count"]) > 0 for c in ["weight", "psu", "strata"])
        full_metadata_triplet = all(int(row[f"{c}_metadata_evidence"]) == 1 for c in ["weight", "psu", "strata"])
        weight_only = int(row["weight_cleaned_variable_count"]) > 0
        if full_cleaned_triplet and full_metadata_triplet:
            decision = "full_design_candidate_modelable_after_manual_codebook_confirmation"
        elif weight_only:
            decision = "weight_only_sensitivity_modelable_no_psu_or_strata"
        elif full_metadata_triplet:
            decision = "metadata_mentions_only_no_cleaned_design_variables"
        else:
            decision = "not_modelable_from_current_cleaned_files"
        row["full_cleaned_triplet_available"] = int(full_cleaned_triplet)
        row["full_metadata_triplet_available"] = int(full_metadata_triplet)
        row["phase61_analysis_decision"] = decision
        rows.append(row)
    return pd.DataFrame(rows)


def read_harmonized_lfo_model_frame() -> pd.DataFrame:
    lfo = keyify(read_csv("phase32_decoupled_lfo_participant_screen.csv", dtype={"participant_id": str, "wave": str}))
    cov = keyify(read_csv("phase13_covariate_participant_screen.csv", dtype={"participant_id": str, "wave": str}))
    cov_keep = cov[
        [
            "cohort",
            "participant_id",
            "wave",
            "cov_education_raw",
            "cov_marital_status_raw",
            "cov_smoking_raw",
            "cov_drinking_raw",
        ]
    ].copy()
    df = lfo.merge(cov_keep, on=["cohort", "participant_id", "wave"], how="left")
    df = coerce_numeric(
        df,
        [
            "age",
            "functional_deterioration_ge_0_5sd",
            "functional_deterioration_available",
            "lfo_assignment_available",
            "lfo_profile_class",
            *LFO_PROFILE_COLS,
            "cov_education_raw",
            "cov_marital_status_raw",
            "cov_smoking_raw",
            "cov_drinking_raw",
        ],
    )
    df = df[
        df["functional_deterioration_available"].eq(1)
        & df["lfo_assignment_available"].eq(1)
        & df["cohort"].isin(["CHARLS", "ELSA", "HRS", "MHAS", "SHARE", "KLoSA"])
    ].copy()
    df["lfo_profile_class"] = df["lfo_profile_class"].astype("Int64")
    return df


def read_clean_design_frame(cohort: str, variables: list[str]) -> pd.DataFrame:
    path = find_clean_csv_from_audit(cohort)
    if path is None:
        return pd.DataFrame()
    config = {
        "CHARLS": ("ID", "wave"),
        "ELSA": ("idauniqc", "wave"),
        "HRS": ("hhidpn", "wave"),
        "KLoSA": ("pid", "wave"),
        "LASI": ("prim_key", ""),
        "MHAS": ("rahhidnp", "wave"),
        "SHARE": ("mergeid", "wave"),
    }[cohort]
    id_col, wave_col = config
    header = read_header_map(path)
    wanted = [id_col, *variables]
    if wave_col:
        wanted.append(wave_col)
    available = {var: header[var] for var in wanted if var in header}
    if id_col not in available:
        return pd.DataFrame()
    read = pd.read_csv(
        path,
        usecols=list(available.values()),
        dtype=str,
        encoding="utf-8-sig",
        low_memory=False,
    ).rename(columns={raw: var for var, raw in available.items()})
    out = pd.DataFrame(
        {
            "cohort": cohort,
            "participant_id": read[id_col].astype(str),
            "wave": read[wave_col].map(normalize_wave) if wave_col and wave_col in read.columns else "all_rows_no_wave",
        }
    )
    for variable in variables:
        if variable in read.columns:
            out[variable] = pd.to_numeric(read[variable], errors="coerce")
    return out


def fit_logistic(formula: str, data: pd.DataFrame, weight_col: str | None = None):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        if weight_col is None:
            return smf.glm(formula=formula, data=data, family=sm.families.Binomial()).fit(maxiter=200)
        return smf.glm(
            formula=formula,
            data=data,
            family=sm.families.Binomial(),
            freq_weights=data[weight_col].astype(float),
        ).fit(maxiter=200)


def weighted_auc(y: pd.Series, pred: np.ndarray, weight: pd.Series | None) -> float:
    if y.nunique() < 2:
        return np.nan
    if weight is None:
        return float(roc_auc_score(y.astype(int), pred))
    return float(roc_auc_score(y.astype(int), pred, sample_weight=weight.astype(float)))


def summarize_fit_terms(fit, metadata: dict[str, object]) -> pd.DataFrame:
    conf = fit.conf_int()
    rows = []
    for term, estimate in fit.params.items():
        lo, hi = conf.loc[term]
        rows.append(
            {
                **metadata,
                "term": term,
                "log_or": float(estimate),
                "or": safe_exp(float(estimate)),
                "ci_low": safe_exp(float(lo)),
                "ci_high": safe_exp(float(hi)),
                "p_value": float(fit.pvalues[term]),
            }
        )
    return pd.DataFrame(rows)


def run_phase61_weight_only_sensitivity(status: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    lfo = read_harmonized_lfo_model_frame()
    metrics_rows = []
    term_frames = []
    class_rows = []
    skipped = []

    profile_formula = (
        "functional_deterioration_ge_0_5sd ~ age + cov_education_raw + cov_marital_status_raw "
        "+ cov_smoking_raw + cov_drinking_raw + C(lfo_profile_class, Treatment(reference=1))"
    )
    continuous_formula = (
        "functional_deterioration_ge_0_5sd ~ age + cov_education_raw + cov_marital_status_raw "
        "+ cov_smoking_raw + cov_drinking_raw + cognitive_score + affective_score + cardiometabolic_chronic_score"
    )
    required_common = [
        "functional_deterioration_ge_0_5sd",
        "age",
        "cov_education_raw",
        "cov_marital_status_raw",
        "cov_smoking_raw",
        "cov_drinking_raw",
    ]

    for _, row in status.iterrows():
        cohort = str(row["cohort"])
        weight_vars = [v for v in str(row.get("weight_cleaned_variables", "")).split(";") if v]
        if not weight_vars:
            skipped.append(
                {
                    "cohort": cohort,
                    "reason": "no cleaned weight variable available for model frame",
                    "phase61_analysis_decision": row["phase61_analysis_decision"],
                }
            )
            continue
        weight_var = weight_vars[0]
        design = read_clean_design_frame(cohort, [weight_var])
        if design.empty or weight_var not in design.columns:
            skipped.append(
                {
                    "cohort": cohort,
                    "reason": f"cleaned weight variable {weight_var} could not be read",
                    "phase61_analysis_decision": row["phase61_analysis_decision"],
                }
            )
            continue

        base = lfo[lfo["cohort"].eq(cohort)].copy()
        merged = base.merge(design, on=["cohort", "participant_id", "wave"], how="left")
        merged = merged.rename(columns={weight_var: "survey_weight_raw"})
        merged["survey_weight_raw"] = pd.to_numeric(merged["survey_weight_raw"], errors="coerce")
        merged = merged[merged["survey_weight_raw"].notna() & (merged["survey_weight_raw"] > 0)].copy()
        if merged.empty:
            skipped.append(
                {
                    "cohort": cohort,
                    "reason": f"no positive nonmissing {weight_var} values after merge",
                    "phase61_analysis_decision": row["phase61_analysis_decision"],
                }
            )
            continue
        merged["survey_weight_norm"] = merged["survey_weight_raw"] / merged["survey_weight_raw"].mean()
        lo, hi = merged["survey_weight_norm"].quantile([0.01, 0.99])
        merged["survey_weight_norm_trim99"] = merged["survey_weight_norm"].clip(lower=float(lo), upper=float(hi))
        merged["survey_weight_norm_trim99"] = merged["survey_weight_norm_trim99"] / merged["survey_weight_norm_trim99"].mean()

        for model_type, formula, extra_required in [
            ("lfo_profile", profile_formula, ["lfo_profile_class"]),
            ("continuous_three_domain", continuous_formula, LFO_PROFILE_COLS),
        ]:
            data = merged.dropna(subset=required_common + extra_required + ["survey_weight_norm", "survey_weight_norm_trim99"]).copy()
            data["functional_deterioration_ge_0_5sd"] = data["functional_deterioration_ge_0_5sd"].astype(int)
            if "lfo_profile_class" in data.columns:
                data["lfo_profile_class"] = data["lfo_profile_class"].astype(int)
            if data.empty or data["functional_deterioration_ge_0_5sd"].nunique() < 2:
                skipped.append({"cohort": cohort, "reason": f"{model_type}: no usable binary model frame"})
                continue
            y = data["functional_deterioration_ge_0_5sd"]
            for weighting in ["unweighted", "weight_only_norm", "weight_only_trim99"]:
                weight_col = None if weighting == "unweighted" else (
                    "survey_weight_norm" if weighting == "weight_only_norm" else "survey_weight_norm_trim99"
                )
                fit = fit_logistic(formula, data, weight_col=weight_col)
                pred = fit.predict(data)
                weight = None if weight_col is None else data[weight_col]
                metadata = {
                    "cohort": cohort,
                    "model_type": model_type,
                    "weighting": weighting,
                    "weight_variable": weight_var,
                    "n": int(len(data)),
                    "events": int(y.sum()),
                    "event_pct_unweighted": float(y.mean() * 100.0),
                    "event_pct_weighted": weighted_mean(y, weight) * 100.0 if weight is not None else float(y.mean() * 100.0),
                    "auc": weighted_auc(y, pred, weight),
                    "aic_or_pseudo_aic": float(fit.aic),
                    "weight_mean": float(data["survey_weight_raw"].mean()),
                    "weight_p01": float(data["survey_weight_raw"].quantile(0.01)),
                    "weight_p99": float(data["survey_weight_raw"].quantile(0.99)),
                    "design_note": "weight-only normalized GLM; no cleaned PSU/strata available in current model frame",
                }
                metrics_rows.append(metadata)
                term_frames.append(summarize_fit_terms(fit, metadata))

            if model_type == "lfo_profile":
                for cls, g in data.groupby("lfo_profile_class", dropna=False):
                    if pd.isna(cls):
                        continue
                    class_rows.append(
                        {
                            "cohort": cohort,
                            "lfo_profile_class": int(cls),
                            "n": int(len(g)),
                            "events": int(g["functional_deterioration_ge_0_5sd"].sum()),
                            "event_pct_unweighted": float(g["functional_deterioration_ge_0_5sd"].mean() * 100.0),
                            "event_pct_weighted_norm": weighted_mean(
                                g["functional_deterioration_ge_0_5sd"],
                                g["survey_weight_norm"],
                            )
                            * 100.0,
                            "event_pct_weighted_trim99": weighted_mean(
                                g["functional_deterioration_ge_0_5sd"],
                                g["survey_weight_norm_trim99"],
                            )
                            * 100.0,
                            "weight_variable": weight_var,
                            "design_note": "weight-only descriptive class risk; no PSU/strata",
                        }
                    )

    metrics = pd.DataFrame(metrics_rows)
    terms = pd.concat(term_frames, ignore_index=True) if term_frames else pd.DataFrame()
    classes = pd.DataFrame(class_rows)
    skipped_df = pd.DataFrame(skipped)
    return metrics, terms, classes, skipped_df


def class_metrics(labels: np.ndarray) -> tuple[int, float, float]:
    counts = pd.Series(labels).value_counts()
    n = len(labels)
    return int(counts.size), float(counts.min() / n * 100.0), float(counts.max() / n * 100.0)


def order_labels_by_severity(labels: np.ndarray, severity: np.ndarray) -> np.ndarray:
    series = pd.DataFrame({"label": labels, "severity": severity})
    order = series.groupby("label")["severity"].mean().sort_values().index.tolist()
    mapping = {label: idx + 1 for idx, label in enumerate(order)}
    return np.array([mapping[label] for label in labels], dtype=int)


def gmm_covariance_flags(model: GaussianMixture) -> tuple[float, float, int]:
    mins: list[float] = []
    conds: list[float] = []
    cov = model.covariances_
    if model.covariance_type == "full":
        for c in cov:
            eig = np.linalg.eigvalsh(c)
            mn = float(np.min(eig))
            mx = float(np.max(eig))
            mins.append(mn)
            conds.append(float(mx / max(mn, 1e-12)))
    elif model.covariance_type == "tied":
        eig = np.linalg.eigvalsh(cov)
        mn = float(np.min(eig))
        mx = float(np.max(eig))
        mins.append(mn)
        conds.append(float(mx / max(mn, 1e-12)))
    elif model.covariance_type == "diag":
        arr = np.asarray(cov)
        mins.extend(np.min(arr, axis=1).astype(float).tolist())
        conds.extend((np.max(arr, axis=1) / np.maximum(np.min(arr, axis=1), 1e-12)).astype(float).tolist())
    elif model.covariance_type == "spherical":
        arr = np.asarray(cov).astype(float).ravel()
        mins.extend(arr.tolist())
        conds.extend([1.0] * len(arr))
    min_var = float(np.min(mins)) if mins else np.nan
    max_condition = float(np.max(conds)) if conds else np.nan
    flag = int(
        (not pd.isna(min_var) and min_var < MIN_GMM_VAR)
        or (not pd.isna(max_condition) and max_condition > MAX_GMM_CONDITION)
    )
    return min_var, max_condition, flag


def safe_silhouette(x: np.ndarray, labels: np.ndarray, rng: np.random.Generator) -> float:
    if len(np.unique(labels)) < 2 or len(labels) < 10:
        return np.nan
    n = len(labels)
    idx = np.arange(n) if n <= 5000 else np.sort(rng.choice(n, size=5000, replace=False))
    if len(np.unique(labels[idx])) < 2:
        return np.nan
    return float(silhouette_score(x[idx], labels[idx]))


def fit_method(x: np.ndarray, k: int, method: str, seed: int, n_init: int | None = None):
    if method.startswith("gmm_"):
        parts = method.split("_")
        covariance_type = parts[1]
        reg_covar = 1e-6
        if len(parts) > 2 and parts[2].startswith("reg"):
            reg_covar = float(parts[2].replace("reg", ""))
        model = GaussianMixture(
            n_components=k,
            covariance_type=covariance_type,
            reg_covar=reg_covar,
            random_state=seed,
            n_init=n_init or PHASE62_FULL_FIT_GMM_N_INIT,
            max_iter=300,
        )
        raw = model.fit_predict(x)
        labels = order_labels_by_severity(raw, x.mean(axis=1))
        return labels, model
    if method == "kmeans":
        model = KMeans(
            n_clusters=k,
            random_state=seed,
            n_init=n_init or PHASE62_FULL_FIT_KMEANS_N_INIT,
            max_iter=300,
        )
        raw = model.fit_predict(x)
        labels = order_labels_by_severity(raw, x.mean(axis=1))
        return labels, model
    if method == "severity_quantile_k":
        severity = pd.Series(x.mean(axis=1))
        labels = pd.qcut(severity.rank(method="first"), q=k, labels=False, duplicates="drop").astype(int).to_numpy() + 1
        return labels, None
    raise ValueError(f"Unknown method: {method}")


def predict_method(model, x: np.ndarray, method: str, sample_x: np.ndarray | None = None, k: int | None = None) -> np.ndarray:
    if method.startswith("gmm_"):
        raw = model.predict(x)
        return order_labels_by_severity(raw, x.mean(axis=1))
    if method == "kmeans":
        raw = model.predict(x)
        return order_labels_by_severity(raw, x.mean(axis=1))
    if method == "severity_quantile_k":
        assert sample_x is not None and k is not None
        sample_sev = sample_x.mean(axis=1)
        cuts = np.quantile(sample_sev, np.linspace(0, 1, k + 1)[1:-1])
        full_sev = x.mean(axis=1)
        return np.searchsorted(cuts, full_sev, side="right") + 1
    raise ValueError(method)


def normalized_entropy(probs: np.ndarray) -> float:
    clipped = np.clip(probs, 1e-12, 1.0)
    entropy = -np.mean(np.sum(clipped * np.log(clipped), axis=1))
    return float(1.0 - entropy / np.log(probs.shape[1]))


def phase62_methods_for_k(k: int) -> list[str]:
    return [
        "gmm_diag",
        "kmeans",
        "severity_quantile_k",
    ]


def run_phase62_stability_gate() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    assignments = keyify(read_csv("phase4_best_model_assignments.csv", dtype={"participant_id": str, "wave": str}))
    assignments = coerce_numeric(assignments, [*DOMAIN_COLS, "endotype_class", "n_classes"])
    rng = np.random.default_rng(RANDOM_SEED)
    metrics_rows = []
    boot_rows = []
    profile_rows = []

    for cohort in COHORT_ORDER:
        g = assignments[assignments["cohort"].eq(cohort)].dropna(subset=DOMAIN_COLS + ["endotype_class", "n_classes"]).copy()
        if g.empty:
            continue
        print(f"Phase62 cohort {cohort}: n={len(g)}", flush=True)
        x = g[DOMAIN_COLS].astype(float).to_numpy()
        severity = x.mean(axis=1)
        selected = g["endotype_class"].astype(int).to_numpy()
        k = int(g["n_classes"].iloc[0])
        analysis_set = str(g["analysis_set"].iloc[0])
        analysis_tier = str(g["analysis_tier"].iloc[0])
        wave = str(g["wave"].iloc[0])

        all_methods = ["selected_full_gmm_reference", *phase62_methods_for_k(k)]
        method_labels: dict[str, np.ndarray] = {"selected_full_gmm_reference": selected}

        for method in all_methods:
            print(f"  method {method}", flush=True)
            model = None
            if method == "selected_full_gmm_reference":
                labels = selected
                min_var = np.nan
                max_condition = np.nan
                cov_flag = np.nan
                bic = np.nan
                aic = np.nan
                entropy = np.nan
                converged = np.nan
            else:
                try:
                    labels, model = fit_method(x, k, method, RANDOM_SEED)
                    if method.startswith("gmm_") and isinstance(model, GaussianMixture):
                        min_var, max_condition, cov_flag = gmm_covariance_flags(model)
                        bic = float(model.bic(x))
                        aic = float(model.aic(x))
                        entropy = normalized_entropy(model.predict_proba(x))
                        converged = int(model.converged_)
                    else:
                        min_var = np.nan
                        max_condition = np.nan
                        cov_flag = 0
                        bic = np.nan
                        aic = np.nan
                        entropy = np.nan
                        converged = 1
                    method_labels[method] = labels
                except Exception as exc:
                    metrics_rows.append(
                        {
                            "analysis_set": analysis_set,
                            "analysis_tier": analysis_tier,
                            "cohort": cohort,
                            "wave": wave,
                            "method": method,
                            "target_classes": k,
                            "fit_status": f"failed: {type(exc).__name__}: {exc}",
                        }
                    )
                    continue

            observed_classes, min_pct, max_pct = class_metrics(labels)
            metrics_rows.append(
                {
                    "analysis_set": analysis_set,
                    "analysis_tier": analysis_tier,
                    "cohort": cohort,
                    "wave": wave,
                    "method": method,
                    "target_classes": k,
                    "n": len(g),
                    "observed_classes": observed_classes,
                    "min_class_pct": min_pct,
                    "max_class_pct": max_pct,
                    "ari_vs_selected_full_gmm": float(adjusted_rand_score(selected, labels)),
                    "silhouette_sample": safe_silhouette(x, labels, rng),
                    "bic": bic,
                    "aic": aic,
                    "entropy": entropy,
                    "converged": converged,
                    "min_covariance_variance_or_eigen": min_var,
                    "max_covariance_condition": max_condition,
                    "covariance_guardrail_flag": cov_flag,
                    "fit_status": "ok",
                }
            )

            prof = pd.DataFrame(x, columns=DOMAIN_COLS)
            prof["class"] = labels
            for cls, pg in prof.groupby("class", dropna=False):
                profile_rows.append(
                    {
                        "analysis_set": analysis_set,
                        "analysis_tier": analysis_tier,
                        "cohort": cohort,
                        "wave": wave,
                        "method": method,
                        "class": int(cls),
                        "n": int(len(pg)),
                        "class_pct": float(len(pg) / len(prof) * 100.0),
                        **{f"mean_{col}": float(pg[col].mean()) for col in DOMAIN_COLS},
                    }
                )

            if method == "selected_full_gmm_reference":
                continue
            if method not in PHASE62_BOOTSTRAP_METHODS:
                continue
            ref_labels = labels
            for rep in range(1, PHASE62_BOOTSTRAP_REPLICATES + 1):
                sample_size = min(len(x), PHASE62_BOOTSTRAP_MAX_N)
                idx = rng.integers(0, len(x), size=sample_size)
                x_sample = x[idx]
                try:
                    boot_n_init = (
                        PHASE62_BOOTSTRAP_GMM_N_INIT
                        if method.startswith("gmm_")
                        else PHASE62_BOOTSTRAP_KMEANS_N_INIT
                    )
                    sample_labels, sample_model = fit_method(x_sample, k, method, RANDOM_SEED + rep, n_init=boot_n_init)
                    pred_labels = predict_method(sample_model, x, method, sample_x=x_sample, k=k)
                    ari = float(adjusted_rand_score(ref_labels, pred_labels))
                    _, rep_min_pct, _ = class_metrics(pred_labels)
                    rep_converged = int(getattr(sample_model, "converged_", True)) if sample_model is not None else 1
                    skip_reason = ""
                except Exception as exc:
                    ari = np.nan
                    rep_min_pct = np.nan
                    rep_converged = 0
                    skip_reason = f"failed: {type(exc).__name__}: {exc}"
                boot_rows.append(
                    {
                        "analysis_set": analysis_set,
                        "analysis_tier": analysis_tier,
                        "cohort": cohort,
                        "wave": wave,
                        "method": method,
                        "target_classes": k,
                        "replicate_id": rep,
                        "adjusted_rand_index_vs_full_fit": ari,
                        "replicate_min_class_pct_on_full_data": rep_min_pct,
                        "replicate_converged": rep_converged,
                        "skip_reason": skip_reason,
                    }
                )

    metrics = pd.DataFrame(metrics_rows)
    bootstrap = pd.DataFrame(boot_rows)
    profiles = pd.DataFrame(profile_rows)
    summary_rows = []
    if not bootstrap.empty:
        grouped = bootstrap.groupby(["analysis_set", "analysis_tier", "cohort", "wave", "method", "target_classes"], dropna=False)
        for keys, group in grouped:
            analysis_set, analysis_tier, cohort, wave, method, target_classes = keys
            ari = pd.to_numeric(group["adjusted_rand_index_vs_full_fit"], errors="coerce")
            rep_min = pd.to_numeric(group["replicate_min_class_pct_on_full_data"], errors="coerce")
            conv = pd.to_numeric(group["replicate_converged"], errors="coerce")
            m = metrics[
                metrics["cohort"].eq(cohort)
                & metrics["method"].eq(method)
                & metrics["wave"].astype(str).eq(str(wave))
            ].iloc[0]
            cov_flag = int(m.get("covariance_guardrail_flag", 0) or 0)
            min_class = float(m.get("min_class_pct", np.nan))
            median_ari = float(ari.median()) if ari.notna().any() else np.nan
            p10_ari = float(ari.quantile(0.10)) if ari.notna().any() else np.nan
            converged_pct = float(conv.mean() * 100.0) if conv.notna().any() else np.nan
            pass_gate = (
                converged_pct >= 95.0
                and not pd.isna(median_ari)
                and median_ari >= STABLE_MEDIAN_ARI
                and p10_ari >= STABLE_P10_ARI
                and min_class >= MIN_CLASS_PCT_GATE
                and cov_flag == 0
            )
            if pass_gate and method == "severity_quantile_k":
                conclusion = "stable_severity_strata_not_latent_profile"
            elif pass_gate:
                conclusion = "passes_phase62_model_based_stability_gate"
            elif cov_flag == 1:
                conclusion = "fails_covariance_guardrail"
            elif min_class < MIN_CLASS_PCT_GATE:
                conclusion = "fails_min_class_gate"
            elif pd.isna(median_ari) or p10_ari < STABLE_P10_ARI:
                conclusion = "fails_bootstrap_ari_gate"
            else:
                conclusion = "fails_stability_gate"
            summary_rows.append(
                {
                    "analysis_set": analysis_set,
                    "analysis_tier": analysis_tier,
                    "cohort": cohort,
                    "wave": wave,
                    "method": method,
                    "target_classes": int(target_classes),
                    "n": int(m.get("n", np.nan)),
                    "observed_classes": int(m.get("observed_classes", np.nan)),
                    "min_class_pct": min_class,
                    "ari_vs_selected_full_gmm": float(m.get("ari_vs_selected_full_gmm", np.nan)),
                    "silhouette_sample": float(m.get("silhouette_sample", np.nan)),
                    "bootstrap_replicates": int(len(group)),
                    "bootstrap_converged_pct": converged_pct,
                    "median_ari_vs_full_fit": median_ari,
                    "p10_ari_vs_full_fit": p10_ari,
                    "min_ari_vs_full_fit": float(ari.min()) if ari.notna().any() else np.nan,
                    "median_replicate_min_class_pct": float(rep_min.median()) if rep_min.notna().any() else np.nan,
                    "covariance_guardrail_flag": cov_flag,
                    "phase62_gate_pass": int(pass_gate),
                    "phase62_gate_conclusion": conclusion,
                }
            )
    summary = pd.DataFrame(summary_rows)
    cohort_summary_rows = []
    for cohort in COHORT_ORDER:
        g = summary[summary["cohort"].eq(cohort)].copy()
        model_pass = g[
            g["phase62_gate_conclusion"].eq("passes_phase62_model_based_stability_gate")
            & ~g["method"].eq("severity_quantile_k")
        ].copy()
        severity_pass = g[g["phase62_gate_conclusion"].eq("stable_severity_strata_not_latent_profile")].copy()
        if not model_pass.empty:
            best = model_pass.sort_values(
                ["p10_ari_vs_full_fit", "median_ari_vs_full_fit", "ari_vs_selected_full_gmm"],
                ascending=False,
            ).iloc[0]
            decision = "model_based_profile_family_supported_as_descriptive_sensitivity"
            best_method = best["method"]
        elif not severity_pass.empty:
            best = severity_pass.sort_values(["p10_ari_vs_full_fit", "median_ari_vs_full_fit"], ascending=False).iloc[0]
            decision = "only_severity_strata_pass_stability_gate"
            best_method = best["method"]
        else:
            best = g.sort_values(["p10_ari_vs_full_fit", "median_ari_vs_full_fit"], ascending=False).iloc[0] if not g.empty else None
            decision = "no_phase62_candidate_passed_stability_gate"
            best_method = best["method"] if best is not None else ""
        cohort_summary_rows.append(
            {
                "cohort": cohort,
                "phase62_decision": decision,
                "best_available_method": best_method,
                "best_p10_ari": float(best["p10_ari_vs_full_fit"]) if best is not None else np.nan,
                "best_median_ari": float(best["median_ari_vs_full_fit"]) if best is not None else np.nan,
                "best_ari_vs_selected_full_gmm": float(best["ari_vs_selected_full_gmm"]) if best is not None else np.nan,
                "endotype_claim_allowed": "no",
                "recommended_claim": "descriptive sensitivity only; continuous domain scores remain primary for modelling",
            }
        )
    cohort_summary = pd.DataFrame(cohort_summary_rows)
    return metrics, bootstrap, summary, profiles, cohort_summary


def autosize(ws) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
        ws.column_dimensions[letter].width = max(10, min(max_len + 2, 60))


def write_dataframe_sheet(wb, sheet_name: str, df: pd.DataFrame) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            ws.cell(row=i, column=j, value=value)
    ws.freeze_panes = "A2"
    autosize(ws)


def update_sheet_index(wb) -> None:
    descriptions = {
        "phase61_corrected_status": "Corrected cleaned-file and metadata survey-design availability status.",
        "phase61_weighted_metrics": "Weight-only LFO sensitivity model metrics where cleaned weights were available.",
        "phase61_weighted_terms": "Weight-only LFO sensitivity model coefficients.",
        "phase61_weighted_class_risks": "Unweighted and weight-only class risks for LFO profiles.",
        "phase61_skipped": "Cohorts skipped from Phase61 modelling and reasons.",
        "phase62_cohort_decision": "Cohort-level stability-gated profile-family decision.",
        "phase62_gate_summary": "Bootstrap stability-gate summary for alternative algorithms.",
        "phase62_algorithm_metrics": "Full-data alternative algorithm metrics.",
        "phase62_bootstrap": "Bootstrap replicate-level stability outputs for alternative algorithms.",
        "phase62_class_profiles": "Class-level domain profile means for alternative algorithms.",
    }
    if "sheet_index" in wb.sheetnames:
        del wb["sheet_index"]
    ws = wb.create_sheet("sheet_index", 0)
    ws.append(["sheet_name", "description"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in wb.sheetnames:
        if sheet == "sheet_index":
            continue
        ws.append([sheet, descriptions.get(sheet, "Workbook sheet used by the manuscript package.")])
    autosize(ws)


def update_submission_workbooks(
    phase61_outputs: dict[str, pd.DataFrame],
    phase62_outputs: dict[str, pd.DataFrame],
) -> None:
    af5 = PKG / "additional_file_5_survey_design_and_sex_comparator_audits.xlsx"
    wb5 = load_workbook(af5)
    write_dataframe_sheet(wb5, "phase61_corrected_status", phase61_outputs["status"])
    write_dataframe_sheet(wb5, "phase61_weighted_metrics", phase61_outputs["metrics"])
    write_dataframe_sheet(wb5, "phase61_weighted_terms", phase61_outputs["terms"])
    write_dataframe_sheet(wb5, "phase61_weighted_class_risks", phase61_outputs["classes"])
    write_dataframe_sheet(wb5, "phase61_skipped", phase61_outputs["skipped"])
    update_sheet_index(wb5)
    wb5.save(af5)

    af2 = PKG / "additional_file_2_profile_stability_and_descriptive_profiles.xlsx"
    wb2 = load_workbook(af2)
    write_dataframe_sheet(wb2, "phase62_cohort_decision", phase62_outputs["cohort_summary"])
    write_dataframe_sheet(wb2, "phase62_gate_summary", phase62_outputs["summary"])
    write_dataframe_sheet(wb2, "phase62_algorithm_metrics", phase62_outputs["metrics"])
    write_dataframe_sheet(wb2, "phase62_bootstrap", phase62_outputs["bootstrap"])
    write_dataframe_sheet(wb2, "phase62_class_profiles", phase62_outputs["profiles"])
    update_sheet_index(wb2)
    wb2.save(af2)


def markdown_table(df: pd.DataFrame, max_rows: int | None = None) -> str:
    if df.empty:
        return "No rows."
    table = df.copy()
    if max_rows is not None and len(table) > max_rows:
        table = table.head(max_rows).copy()
    columns = [str(c) for c in table.columns]
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join(["---"] * len(columns)) + " |",
    ]
    for row in table.itertuples(index=False):
        values = []
        for value in row:
            if pd.isna(value):
                values.append("")
            elif isinstance(value, float):
                values.append(f"{value:.4g}")
            else:
                values.append(str(value).replace("|", "/"))
        lines.append("| " + " | ".join(values) + " |")
    if max_rows is not None and len(df) > max_rows:
        lines.append(f"| ... | {len(df) - max_rows} additional rows omitted from report preview. |" + " |" * max(0, len(columns) - 2))
    return "\n".join(lines)


def write_phase61_report(status: pd.DataFrame, metrics: pd.DataFrame, skipped: pd.DataFrame) -> Path:
    modelled = sorted(metrics["cohort"].dropna().unique().tolist()) if not metrics.empty else []
    lines = [
        "# Phase 61 Survey-Weight Sensitivity",
        "",
        "Date: 2026-06-08",
        "",
        "## Decision",
        "",
        "The corrected cleaned-file audit did not identify a full weight/PSU/strata triplet in the current model frame. Therefore Phase 61 does not support a seven-cohort survey-weighted prevalence or pooled survey-weighted association claim.",
        "",
        f"Weight-only sensitivity models were run for: {', '.join(modelled) if modelled else 'none'}. These models use normalized cleaned weights without PSU/strata and are sensitivity checks only.",
        "",
        "## Weight-only model metrics",
        "",
        markdown_table(metrics) if not metrics.empty else "No weight-only models were fitted.",
        "",
        "## Corrected status",
        "",
        markdown_table(status),
        "",
        "## Skipped cohorts",
        "",
        markdown_table(skipped) if not skipped.empty else "No skipped rows.",
    ]
    path = OUT / "phase61_survey_weighted_sensitivity_report.md"
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_phase62_report(cohort_summary: pd.DataFrame, summary: pd.DataFrame) -> Path:
    model_pass_n = int(cohort_summary["phase62_decision"].eq("model_based_profile_family_supported_as_descriptive_sensitivity").sum())
    severity_only_n = int(cohort_summary["phase62_decision"].eq("only_severity_strata_pass_stability_gate").sum())
    lines = [
        "# Phase 62 Stability-Gated Profile Family Analysis",
        "",
        "Date: 2026-06-08",
        "",
        "## Decision",
        "",
        f"Model-based alternative algorithms passed the pre-specified descriptive stability gate in {model_pass_n} cohorts. Severity-strata-only fallback passed in {severity_only_n} cohorts.",
        "",
        f"The Phase 62 gate used {PHASE62_BOOTSTRAP_REPLICATES} bounded nonparametric bootstrap refits per bootstrapped method, with each refit drawing up to {PHASE62_BOOTSTRAP_MAX_N} original participants with replacement and predicting labels for the full cohort. This is a rapid pre-submission sensitivity analysis, not a definitive external validation study.",
        "",
        "Even where a model-based alternative passes the numerical gate, the result remains a descriptive sensitivity profile because the analysis does not add independent biological mechanism, survey-weighted transportability, or independent hard-endpoint validation.",
        "",
        "Endotype language remains disallowed for all cohorts in this package.",
        "",
        "## Cohort decisions",
        "",
        markdown_table(cohort_summary),
        "",
        "## Gate thresholds",
        "",
        f"- median bootstrap ARI >= {STABLE_MEDIAN_ARI}",
        f"- 10th percentile bootstrap ARI >= {STABLE_P10_ARI}",
        f"- minimum class percentage >= {MIN_CLASS_PCT_GATE}%",
        "- no covariance guardrail flag for model-based GMM alternatives",
        "- at least 95% bootstrap convergence",
        "",
        "## Gate summary",
        "",
        markdown_table(summary, max_rows=80) if not summary.empty else "No gate rows.",
    ]
    path = OUT / "phase62_stability_gated_profile_family_report.md"
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--phase", choices=["all", "phase61", "phase62", "workbooks"], default="all")
    args = parser.parse_args()

    status = metrics61 = terms61 = classes61 = skipped61 = None
    metrics62 = bootstrap62 = summary62 = profiles62 = cohort_summary62 = None
    report61 = report62 = None

    if args.phase in {"all", "phase61"}:
        status = corrected_survey_design_status()
        metrics61, terms61, classes61, skipped61 = run_phase61_weight_only_sensitivity(status)
        write_csv(status, "phase61_survey_design_corrected_status.csv")
        write_csv(metrics61, "phase61_weighted_lfo_model_metrics.csv")
        write_csv(terms61, "phase61_weighted_lfo_model_terms.csv")
        write_csv(classes61, "phase61_weighted_lfo_class_risks.csv")
        write_csv(skipped61, "phase61_weighted_lfo_skipped.csv")
        report61 = write_phase61_report(status, metrics61, skipped61)
        print(report61, flush=True)

    if args.phase in {"all", "phase62"}:
        metrics62, bootstrap62, summary62, profiles62, cohort_summary62 = run_phase62_stability_gate()
        write_csv(metrics62, "phase62_alternative_algorithm_metrics.csv")
        write_csv(bootstrap62, "phase62_alternative_algorithm_bootstrap_replicates.csv")
        write_csv(summary62, "phase62_stability_gate_summary.csv")
        write_csv(profiles62, "phase62_profile_family_class_profiles.csv")
        write_csv(cohort_summary62, "phase62_stability_gated_profile_family_summary.csv")
        report62 = write_phase62_report(cohort_summary62, summary62)
        print(report62, flush=True)

    if args.phase in {"all", "workbooks"}:
        if status is None:
            status = read_csv("phase61_survey_design_corrected_status.csv")
            metrics61 = read_csv("phase61_weighted_lfo_model_metrics.csv")
            terms61 = read_csv("phase61_weighted_lfo_model_terms.csv")
            classes61 = read_csv("phase61_weighted_lfo_class_risks.csv")
            skipped61 = read_csv("phase61_weighted_lfo_skipped.csv")
        if metrics62 is None:
            metrics62 = read_csv("phase62_alternative_algorithm_metrics.csv")
            bootstrap62 = read_csv("phase62_alternative_algorithm_bootstrap_replicates.csv")
            summary62 = read_csv("phase62_stability_gate_summary.csv")
            profiles62 = read_csv("phase62_profile_family_class_profiles.csv")
            cohort_summary62 = read_csv("phase62_stability_gated_profile_family_summary.csv")
        update_submission_workbooks(
            {"status": status, "metrics": metrics61, "terms": terms61, "classes": classes61, "skipped": skipped61},
            {
                "metrics": metrics62,
                "bootstrap": bootstrap62,
                "summary": summary62,
                "profiles": profiles62,
                "cohort_summary": cohort_summary62,
            },
        )
        print(PKG / "additional_file_2_profile_stability_and_descriptive_profiles.xlsx", flush=True)
        print(PKG / "additional_file_5_survey_design_and_sex_comparator_audits.xlsx", flush=True)


if __name__ == "__main__":
    main()
